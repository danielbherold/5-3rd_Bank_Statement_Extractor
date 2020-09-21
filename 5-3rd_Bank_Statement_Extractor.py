# Importing required modules
import PyPDF2
import re
from openpyxl import Workbook, load_workbook
import os

j = k = 0
account_cell = 3

def find_account_number(input_string):
    query = r"(Account Summary - )(\d{10})"
    result = re.findall(query, input_string)
    if result:
        return [["Account", item[1]] for item in result]
    return None

def find_withdraws(input_string):
    query = r"(Withdrawals/Debits)(.*?)(?=Deposits/Credit)"
    result = re.findall(query, input_string)
    if result:
        raw_entries = [item[1] for item in result]
        # print(raw_entries)
        entry_query = r"(\d{2}/\d{2})([0-9,,]+.\d{2})(.*?)(?=\d{2}/\d{2}|$)"
        entries = [["Withdrawals", re.findall(entry_query, item[1])] for item in result]
        if entries:
            return entries
    return None

def find_deposits(input_string):
    query = r"(Deposits/Credit)(.*?)(?=DailyBalanceSummary)"
    result = re.findall(query, input_string)
    if result:
        raw_entries = [item[1] for item in result]
        # print(raw_entries)
        entry_query = r"(\d{2}/\d{2})([0-9,,]+.\d{2})(.*?)(?=\d{2}/\d{2}|$)"
        entries = [["Deposits", re.findall(entry_query, item[1])] for item in result]
        if entries:
            return entries
    return None

def parse_pdf(file_name):
    text = ''
    pdfFileObj = open(file_name,'rb')
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj, strict=False)
    for i in range(pdfReader.numPages):
        pageObj = pdfReader.getPage(i)
        text += pageObj.extractText()
    return text
    pdfFileObj.close()

def collect_data(file_name):
    text = parse_pdf(file_name)
    # print(text)
    account_number = find_account_number(text)
    # print(account_number)
    withdraws = find_withdraws(text)
    # print(withdraws)
    deposits = find_deposits(text)
    # print(deposits)
    return combine_lists(account_number, withdraws, deposits)

def combine_lists(list1, list2, list3):
    for i in range(len(list1)):
        list1[i].append(list2[i])
        list1[i].append(list3[i])
    return list1

def create_Excel_sheet(filename):
    workbook = Workbook()
    sheet = workbook.active

    sheet["A1"] = "Account Number"

    sheet["C1"] = "Withdraws"
    sheet["C2"] = "Date"
    sheet["D2"] = "Amount"
    sheet["E2"] = "Description"

    sheet["G1"] = "Deposits"
    sheet["G2"] = "Date"
    sheet["H2"] = "Amount"
    sheet["I2"] = "Description"

    workbook.save(filename)
    return filename

def edit_Excel_file(filename, data):
    workbook = load_workbook(filename)
    sheet = workbook.active
    global j, k, account_cell
    for i in range(len(data)):
        account_cell += 1
        account_number = data[i][1]
        withdraws = data[i][2][1]
        deposits = data[i][3][1]
        account_cell += max(j,k)
        sheet["A{0}".format(account_cell)] = account_number

        for j in range(len(withdraws)):
            date = withdraws[j][0]
            amount = withdraws[j][1]
            description = withdraws[j][2]
            sheet["C{0}".format(account_cell + j)] = date
            sheet["D{0}".format(account_cell + j)] = amount
            sheet["E{0}".format(account_cell + j)] = description

        for k in range(len(deposits)):
            date = deposits[k][0]
            amount = deposits[k][1]
            description = deposits[k][2]
            sheet["G{0}".format(account_cell + k)] = date
            sheet["H{0}".format(account_cell + k)] = amount
            sheet["I{0}".format(account_cell + k)] = description

        account_cell += 1
    workbook.save(filename)
    return None

def main():
    i = 0
    folder_path = '.'
    workbook_name = "Cash_Flow.xlsx"
    file_type_re = r'\.pdf$'
    print("\nCreating Excel sheet named {0}".format(workbook_name))
    excel_filename = create_Excel_sheet(workbook_name)

    num_files = len([name for name in os.listdir('.') if re.search(file_type_re, name, re.IGNORECASE)])
    print("\nLocating Files in the folder {0}/".format(folder_path))
    print("\nExtracting data from files in the folder {0}/\n".format(folder_path))

    for filename in os.listdir('.'):
        if re.search(file_type_re, filename, re.IGNORECASE):
            i += 1
            print("{0}/{1} Extracting data from {2}".format(i,num_files, filename))
            data = collect_data(filename)
            print("{0}/{1} Writing {2} to Excel File".format(i,num_files, filename))
            edit_Excel_file(workbook_name, data)

if __name__ == "__main__":
    main()
